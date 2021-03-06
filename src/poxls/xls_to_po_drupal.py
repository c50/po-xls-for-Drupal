try:
    from collections import OrderedDict
except ImportError:
    from ordereddict import OrderedDict
import os
import sys
import time
import click
import polib
import openpyxl
from . import ColumnHeaders
try:
    unicode
except NameError:
    unicode = str
import pycountry

def save(output_file, catalog):
    """Save catalog to a PO file.

    This is mostly a stripped down copy of POFile.save so we can save the
    catalog to a file safely created by click.
    """
    output_file.write(unicode(catalog))
    output_file.write("\n") # add final newline


def po_timestamp(filename):
    local = time.localtime(os.stat(filename).st_mtime)
    offset = -(time.altzone if local.tm_isdst else time.timezone)
    return '%s%s%s' % (
        time.strftime('%Y-%m-%d %H:%M', local),
        '-' if offset < 0 else '+',
        time.strftime('%H%M', time.gmtime(abs(offset))))


@click.command()
@click.argument('locale', required=True)
@click.argument('input_file',
        type=click.Path(exists=True, readable=True),
        required=True)
@click.argument('output_file', type=click.File('w', encoding='utf-8'), required=True)
@click.argument('project', required=True)
@click.argument('version', required=True)
def main(locale, input_file, output_file, project, version):
    """
    Convert a XLS(X) file to a .PO file
    """
    language = pycountry.languages.get(alpha_2=locale).name
    click.echo(f'Processing {language} translations...')
    book = openpyxl.load_workbook(input_file)
    catalog = polib.POFile()
    catalog.header = 'Futurium translation for Drupal by CNECT.R3'
    catalog.metata_is_fuzzy = True
    catalog.metadata = OrderedDict()
    catalog.metadata['Project-Id-Version'] = f'{project} ({version})'
    catalog.metadata['POT-Creation-Date'] = po_timestamp(input_file)
    catalog.metadata['PO-Revision-Date'] = 'YYYY-mm-DD HH:MM+ZZZZ'
    catalog.metadata['Content-Type'] = 'text/plain; charset=utf-8'
    catalog.metadata['Content-Transfer-Encoding'] = '8bit'
    catalog.metadata['Language-Team'] = language
    catalog.metadata['MIME-Version'] = '1.0'

    if locale in ['bg', 'da', 'de', 'el', 'en', 'es', 'et', 'fi', 'hu', 'it', 'nl', 'pt', 'sv']:
        plural_form = 'nplurals=2; plural=(n!=1);'
    elif locale in ['fr']:
        plural_form = 'nplurals=2; plural=(n>1);'
    elif locale in ['cs', 'sk']:
        plural_form = 'nplurals=3; plural=((n==1)?(0):(((n>=2)&&(n<=4))?(1):2));'
    elif locale in ['lv']:
        plural_form = 'nplurals=3; plural=(n%10==1 && n%100!=11 ? 0 : n != 0 ? 1 : 2);'
    elif locale in ['ro']:
        plural_form = 'nplurals=3; plural=(n==1 ? 0 : (n==0 || (n%100 > 0 && n%100 < 20)) ? 1 : 2);'
    elif locale in ['lt']:
        plural_form = 'nplurals=3; plural=(n%10==1 && n%100!=11 ? 0 : n%10>=2 && (n%100<10 || n%100>=20) ? 1 : 2);'
    elif locale in ['pl']:
        plural_form = 'nplurals=3; plural=(n==1 ? 0 : n%10>=2 && n%10<=4 && (n%100<10 || n%100>=20) ? 1 : 2);'
    elif locale in ['hr']:
        plural_form = 'nplurals=3; plural=(n%10==1 && n%100!=11 ? 0 : n%10>=2 && n%10<=4 && (n%100<10 || n%100>=20) ? 1 : 2);'
    elif locale in ['sl']:
        plural_form = 'nplurals=4; plural=(n%100==1 ? 0 : n%100==2 ? 1 : n%100==3 || n%100==4 ? 2 : 3);'
    elif locale in ['mt']:
        plural_form = 'nplurals=4; plural=(n==1 ? 0 : n==0 || ( n%100>1 && n%100<11) ? 1 : (n%100>10 && n%100<20 ) ? 2 : 3);'
    elif locale in ['ga']:
        plural_form = 'nplurals=5; plural=n==1 ? 0 : n==2 ? 1 : (n>2 && n<7) ? 2 :(n>6 && n<11) ? 3 : 4;'
    else:
        print(f"Unknown plural form for {locale}")
        sys.exit(1)

    catalog.metadata['Plural-Forms'] = plural_form

    for sheet in book.worksheets:
        if sheet.max_row < 2:
            continue
        row_iterator = sheet.iter_rows()
        headers = [c.value for c in next(row_iterator)]
        headers = dict((b, a) for (a, b) in enumerate(headers))
        msgctxt_column = headers.get(ColumnHeaders.msgctxt)
        msgid_column = headers.get(ColumnHeaders.msgid)
        tcomment_column = headers.get(ColumnHeaders.tcomment)
        comment_column = headers.get(ColumnHeaders.comment)
        msgstr_column = headers.get(locale)
        if msgid_column is None:
            click.echo(u'Could not find a "%s" column' % ColumnHeaders.msgid,
                    err=True)
            continue
        if msgstr_column is None:
            click.echo(u'Could not find a "%s" column' % locale, err=True)
            continue

        rows = list(row_iterator)
        skip_line = False
        for i, row in enumerate(rows):
            if not skip_line:
                row = [c.value for c in row]
                msgid = row[msgid_column]
                if not msgid:
                    continue
                try:
                    entry = polib.POEntry(
                            msgid=msgid
                            )
                    if "1 " in msgid:
                        next_row = [c.value for c in rows[i+1]]
                        if "@count" in next_row[msgid_column]:
                            skip_line = True # don't handle plural twice
                            entry.msgid_plural = next_row[msgid_column]
                            entry.msgstr_plural[0] = row[msgstr_column]
                            if "/" in next_row[msgstr_column]: # handle Slovak case with multiple plural forms
                                plural1, plural2 = next_row[msgstr_column].split("/")
                                if "@count " not in plural1: # missing space
                                    plural1 = plural1.replace("@count", "@count ")
                                entry.msgstr_plural[1] = plural1
                                entry.msgstr_plural[2] = f"@count {plural2}"
                            elif "(" in next_row[msgstr_column]: # handle parentheses
                                plural1 = next_row[msgstr_column].split("(")[0]
                                if "@count " not in plural1: # missing space
                                    plural1 = plural1.replace("@count", "@count ")
                                try:
                                    plural2 = "".join([l for l in next_row[msgstr_column].split()[1] if l.isalpha()])
                                except IndexError:
                                    print(f"Wrong plural format: {next_row[msgstr_column]}")
                                    print("Aborting")
                                    sys.exit()
                                entry.msgstr_plural[1] = plural1
                                entry.msgstr_plural[2] = f"@count {plural2}"
                            else:
                                plural1 = next_row[msgstr_column]
                                if "@count " not in plural1: # missing space
                                    plural1 = plural1.replace("@count", "@count ")
                                entry.msgstr_plural[1] = plural1
                        else:
                            entry.msgstr = row[msgstr_column]    
                    else:
                        entry.msgstr = row[msgstr_column]
                    if msgctxt_column is not None and row[msgctxt_column]:
                        entry.msgctxt = row[msgctxt_column]
                    if tcomment_column:
                        entry.tcomment = row[tcomment_column]
                    if comment_column:
                        entry.comment = row[comment_column]
                    catalog.append(entry)
                except IndexError:
                    click.echo('Row %s is too short' % row, err=True)
            else:
                skip_line = False # reset to process next entry

    if not catalog:
        click.echo('No messages found, aborting', err=True)
        sys.exit(1)

    save(output_file, catalog)


if __name__ == '__main__':
    main()
